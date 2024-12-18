import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_lean_startup_financial_projection(file_path):
    """
    Generate a comprehensive financial projection for S.E.P.T
    focusing on technical infrastructure and MVP development costs.
    """
    # Technical Infrastructure and Cost Projection
    financial_data = {
        "Cost Category": [
            # Technical Infrastructure
            "Cloud Hosting (AWS/Google Cloud)",
            "AI Training Infrastructure (Google TPU)",
            "GPU Compute for Model Training",
            "Data Storage and Management",
            "API and Service Integrations",
            
            # Technical Development Costs
            "Development Tools and Licenses",
            "Software Subscriptions",
            "Security and Compliance",
            
            # AI and Technical Specific Costs
            "NLP Model Training Costs",
            "Graph Neural Network Development",
            "Sentiment Analysis Data Acquisition",
            
            # Operational Technological Expenses
            "Domain and SSL Certificates",
            "Cybersecurity Monitoring",
            
            # Subtotals and Summaries
            "Total Technical Infrastructure Costs",
            "Projected Monthly Burn Rate",
            "Projected Annual Technology Investment"
        ],
        "MVP Development (6 months)": [
            # Technical Infrastructure
            1200,    # Cloud Hosting (modest startup tier)
            5000,    # Google TPU usage for initial model training
            3500,    # Specialized GPU compute resources
            800,     # Data storage and management
            1000,    # Initial API integrations
            
            # Development Tools
            1500,    # Development environment and tools
            600,     # Essential software subscriptions
            750,     # Initial security compliance setup
            
            # AI Specific Costs
            4000,    # Initial NLP model training
            3500,    # GNN development and testing
            2000,    # Sentiment analysis data collection
            
            # Operational Expenses
            150,     # Domain and SSL
            500,     # Basic cybersecurity monitoring
            
            # Calculations
            "=SUM(B2:B13)",
            "=B15/6",
            "=B15*2"
        ],
        "Initial Scaling (Year 1-2)": [
            # Technical Infrastructure
            3600,    # Enhanced cloud hosting
            15000,   # Expanded TPU usage for model refinement
            10000,   # Advanced GPU compute
            2400,    # Increased data management
            3000,    # Comprehensive API integrations
            
            # Development Tools
            3000,    # Advanced development tools
            1800,    # Expanded software subscriptions
            2250,    # Enhanced security compliance
            
            # AI Specific Costs
            12000,   # Advanced NLP model iterations
            10000,   # Complex GNN development
            6000,    # Comprehensive sentiment data
            
            # Operational Expenses
            600,     # Multiple domain management
            1500,    # Advanced cybersecurity
            
            # Calculations
            "=SUM(C2:C13)",
            "=C15/12",
            "=C15"
        ],
        "Market Expansion (Year 3-5)": [
            # Technical Infrastructure
            12000,   # Enterprise-grade cloud hosting
            50000,   # Advanced TPU and AI training
            35000,   # High-performance GPU clusters
            8000,    # Comprehensive data management
            10000,   # Enterprise API ecosystems
            
            # Development Tools
            10000,   # Professional development infrastructure
            6000,    # Comprehensive software ecosystem
            7500,    # Enterprise security compliance
            
            # AI Specific Costs
            40000,   # Advanced AI model development
            35000,   # Sophisticated GNN capabilities
            20000,   # Global sentiment analysis
            
            # Operational Expenses
            2000,    # Multiple global domains
            5000,    # Enterprise cybersecurity
            
            # Calculations
            "=SUM(D2:D13)",
            "=D15/12",
            "=D15"
        ]
    }
    
    # Create DataFrame
    df = pd.DataFrame(financial_data)
    
    # Write to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='S.E.P.T Technical Cost Projection', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['S.E.P.T Technical Cost Projection']
        
        # Custom Formatting
        for row in range(2, len(df) + 2):
            for col in range(2, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0'
                
                # Subtotal and calculation rows special formatting
                if row in [15, 16, 17]:
                    cell.font = Font(bold=True)
        
        # Header Formatting
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 3
    
    print(f"Lean Startup Technical Cost Projection created at {file_path}")
    return file_path

# Example usage
file_path = "/mnt/data/S.E.P.T_Lean_Technical_Projection.xlsx"
result = create_lean_startup_financial_projection(file_path)
print(result)
